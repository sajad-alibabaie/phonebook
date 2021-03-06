from openpyxl import load_workbook
from openpyxl import Workbook
import os.path

from person import person

filepath = 'phbook.xlsx'
numbers = {}


def print_menu():
    print('1. Print Phone Numbers')
    print('2. Add a Phone Number')
    print('3. Remove a Phone Number')
    print('4. Lookup a Phone Number')
    print('5. Lookup a Name by Number')
    print('6. Quit')
	print('press any key to show menu')
    print()


def makeexcel():
    wb = Workbook()
    ws = wb.create_sheet("datashit", 0)
    ws.cell(1, 1).value = "id"
    ws.cell(1, 2).value = "Name"
    ws.cell(1, 3).value = "Number"
    ws.cell(1, 4).value = "Email"
    ws.cell(1, 5).value = "Creat Date"
    wb.save(filepath)


def readexcel():
    wb = load_workbook(filepath)
    return wb


def showpersons():
    print("Telephone Numbers:")
    for x in numbers.keys():
        print("Name: ", x, "\tNumber:", numbers[x][1], "\tEmail:", numbers[x][2], "\tCreate:", numbers[x][3])
    print()



def addperson(ws):
    print("Add Person : ")
    name = input("Name: ")
    phone = input("Number: ")
    email = input("email: ")
    lastEmptyRow = int(len(ws['A'])) + 1
    pr = person(lastEmptyRow, name, phone, email)
    numbers[name] = [lastEmptyRow, pr.pnumber, pr.email, pr.createdate]
    ws.cell(row=lastEmptyRow, column=1).value = pr.id
    ws.cell(row=lastEmptyRow, column=2).value = pr.name
    ws.cell(row=lastEmptyRow, column=3).value = pr.pnumber
    ws.cell(row=lastEmptyRow, column=4).value = pr.email
    ws.cell(row=lastEmptyRow, column=5).value = pr.createdate
    wb.save(filepath)


def findbynumber():
    print("Lookup Name")
    num = input("Number: ")
    find = False
    for name, data in numbers.items():  # for name, age in dictionary.iteritems():  (for Python 2.x)
        if data[1] == num:
            find = True
            print("Name : ", name)
            break
    if (not find):
        print("was not found")


def findbyname():
    print("Lookup Number")
    name = input("Name: ")
    if name in numbers:
        print("The number is", numbers[name][1])
    else:
        print(name, "was not found")


def removebyname(ws):
    print("Remove Name and Number")
    name = input("Name: ")
    if name in numbers:
        ws.delete_rows(numbers[name][1], 1)
        del numbers[name]


    else:
        print(name, "was not found")


if os.path.isfile(filepath):
    wb = readexcel()
    ws = wb.get_sheet_by_name("datashit")
    for i in range(2, len(ws['A']) + 1):
        numbers[ws.cell(i, 2).value] = [ws.cell(i, 1).value, ws.cell(i, 3).value,
                                        ws.cell(i, 4).value, ws.cell(i, 5).value]

else:
    makeexcel()
    wb = readexcel()
    ws = wb.get_sheet_by_name("datashit")

menu_choice = 0
print_menu()
while menu_choice != 6:
    menu_choice = input("Type in a number (1-6): ")

    if menu_choice == "1":
        showpersons()

    elif menu_choice == "2":
        addperson(ws)

    elif menu_choice == "3":
        removebyname(ws)

    elif menu_choice == "4":
        findbyname()

    elif menu_choice == "5":
        findbynumber()

    elif menu_choice == "6":
        break

    else:
        print_menu()
